#!/usr/bin/env python3
"""
电影小镇数据查询工具 V3（三维度综合版）
功能：统一查询往年数据、计划目标、实际完成数据

用法:
  python3 scripts/query_data_v3.py --validate   # 校验数据文件
  python3 scripts/query_data_v3.py --month 2     # 查询2月三维度数据
  python3 scripts/query_data_v3.py --q1         # 查询Q1汇总
  python3 scripts/query_data_v3.py --holiday 春节  # 查询春节数据
"""

import argparse
import openpyxl
from datetime import datetime
from pathlib import Path

WORKSPACE = Path("/Users/tianjinzhan/.openclaw/workspace")
DESKTOP = Path("/Users/tianjinzhan/Desktop")

PATHS = {
    "往年数据": WORKSPACE / "2023-2025年门票销售及客流统计数据表.xlsx",
    "26年计划": WORKSPACE / "电影小镇2026年经营计划.md",
    "26年实际": DESKTOP / "2026年电影小镇实际客流.xlsx",
}


def validate():
    """校验数据文件"""
    print("=" * 60)
    print("数据文件校验")
    print("=" * 60)
    all_ok = True
    for name, path in PATHS.items():
        exists = path.exists()
        status = "✅" if exists else "❌"
        print(f"{status} {name}: {path}")
        if not exists:
            all_ok = False
    return all_ok


def load_historical_month(month):
    """加载往年指定月份数据"""
    wb = openpyxl.load_workbook(PATHS["往年数据"])
    
    result = {}
    for year in [2023, 2024, 2025]:
        try:
            ws = wb[f"{year}年"]
        except KeyError:
            continue
        
        # 找日期列
        dates = []
        for col in range(3, 50):  # 最多50列
            cell = ws.cell(1, col).value
            if cell and isinstance(cell, datetime):
                dates.append((col, cell))
        
        # 找合计行 (在第1列)
        total_row = income_row = None
        for row in range(1, 12):
            cell = ws.cell(row, 1).value
            if cell == "门票人数合计（单位：人）":
                total_row = row
            elif cell == "门票收入金额（单位：元）":
                income_row = row
        
        if not total_row or not income_row:
            continue
        
        # 汇总当月数据
        客流 = 0
        收入 = 0
        for col in range(3, 70):  # 扩大搜索范围
            cell = ws.cell(1, col).value
            if cell and isinstance(cell, datetime) and cell.month == month and cell.year == year:
                客流 += ws.cell(total_row, col).value or 0
                收入 += ws.cell(income_row, col).value or 0
        
        if 客流 or 收入:
            result[year] = {"客流": 客流, "收入": 收入}
    
    return result


def load_2026_actual_month(month):
    """加载2026年实际数据"""
    if not PATHS["26年实际"].exists():
        return None
    
    wb = openpyxl.load_workbook(PATHS["26年实际"])
    ws = wb.active
    
    # 找日期列
    dates = []
    for col in range(3, 100):
        cell = ws.cell(1, col).value
        if cell and isinstance(cell, datetime) and cell.year == 2026 and cell.month == month:
            dates.append(col)
    
    if not dates:
        return None
    
    客流 = sum(ws.cell(9, col).value or 0 for col in dates)
    收入 = sum(ws.cell(10, col).value or 0 for col in dates)
    
    return {"客流": 客流, "收入": 收入} if 客流 or 收入 else None


# 计划目标
PLANS = {
    1: {"客流": 380000, "收入": 35000000},
    2: {"客流": 180000, "收入": 18000000},
    3: {"客流": 80000, "收入": 7000000},
    4: {"客流": 80000, "收入": 8000000},
    5: {"客流": 180000, "收入": 18000000},
    6: {"客流": 600000, "收入": 6000000},
    7: {"客流": 850000, "收入": 7500000},
    8: {"客流": 1200000, "收入": 10000000},
    9: {"客流": 600000, "收入": 6000000},
    10: {"客流": 2000000, "收入": 20000000},
    11: {"客流": 350000, "收入": 3000000},
    12: {"客流": 500000, "收入": 4500000},
}


def query_month(month):
    """查询单月三维度数据"""
    print("=" * 60)
    print(f"【{month}月】数据查询")
    print("=" * 60)
    
    # 1. 往年数据
    print("\n📊 往年数据:")
    hist = load_historical_month(month)
    for year in sorted(hist.keys()):
        d = hist[year]
        print(f"  {year}年: 客流 {d['客流']:>8,} | 收入 {d['收入']/10000:>8.1f}万")
    
    # 2. 计划目标
    if month in PLANS:
        p = PLANS[month]
        print(f"\n📋 26年计划: 客流 {p['客流']:>8,} | 收入 {p['收入']/10000:>8.1f}万")
    
    # 3. 实际完成
    actual = load_2026_actual_month(month)
    if actual and (actual['客流'] or actual['收入']):
        print(f"\n✅ 26年实际: 客流 {actual['客流']:>8,} | 收入 {actual['收入']/10000:>8.1f}万")
        
        if month in PLANS:
            diff_客流 = actual['客流'] - PLANS[month]['客流']
            diff_收入 = actual['收入'] - PLANS[month]['收入']
            print(f"   vs计划: 客流 {diff_客流:+} | 收入 {diff_收入/10000:+.1f}万")
            
            if 2024 in hist:
                diff_客流24 = actual['客流'] - hist[2024]['客流']
                diff_收入24 = actual['收入'] - hist[2024]['收入']
                print(f"   vs 2024: 客流 {diff_客流24:+} | 收入 {diff_收入24/10000:+.1f}万")
    else:
        print(f"\n⚠️ 26年实际: 数据未录入")


def query_q1():
    """Q1汇总"""
    print("=" * 60)
    print("【Q1季度汇总】")
    print("=" * 60)
    
    hist_total = {2023: 0, 2024: 0, 2025: 0}
    actual_total = {"客流": 0, "收入": 0}
    plan_total = {"客流": 0, "收入": 0}
    
    for month in [1, 2, 3]:
        hist = load_historical_month(month)
        for y in hist:
            hist_total[y] += hist[y]['客流']
        
        actual = load_2026_actual_month(month)
        if actual:
            actual_total['客流'] += actual['客流']
            actual_total['收入'] += actual['收入']
        
        plan_total['客流'] += PLANS[month]['客流']
        plan_total['收入'] += PLANS[month]['收入']
    
    print("\n📊 往年Q1:")
    for y in sorted(hist_total.keys()):
        print(f"  {y}年: 客流 {hist_total[y]:>8,}")
    
    print(f"\n📋 26年计划: 客流 {plan_total['客流']:,}")
    print(f"✅ 26年实际: 客流 {actual_total['客流']:,}")
    
    if actual_total['客流']:
        print(f"   vs计划: {actual_total['客流'] - plan_total['客流']:+}")
        if 2024 in hist_total:
            print(f"   vs 2024: {actual_total['客流'] - hist_total[2024]:+}")


def main():
    parser = argparse.ArgumentParser(description="电影小镇三维度数据查询")
    parser.add_argument("--validate", action="store_true", help="校验数据文件")
    parser.add_argument("--month", type=int, help="查询月份(1-12)")
    parser.add_argument("--q1", action="store_true", help="查询Q1汇总")
    
    args = parser.parse_args()
    
    if args.validate or (not args.month and not args.q1):
        validate()
    
    if args.month:
        query_month(args.month)
    
    if args.q1:
        query_q1()


if __name__ == "__main__":
    main()
