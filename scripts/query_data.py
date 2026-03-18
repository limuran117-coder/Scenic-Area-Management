#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
电影小镇历史数据查询工具
使用方法: python3 query_data.py --type qingming --year 2024
         python3 query_data.py --type wuyi --year all
         python3 query_data.py --type month --month 4 --year all
"""

import openpyxl
from datetime import datetime
import argparse
import sys

# 数据文件路径
DATA_FILE = '2023-2025年门票销售及客流统计数据表.xlsx'

def load_workbook():
    """加载数据文件"""
    try:
        wb = openpyxl.load_workbook(DATA_FILE)
        return wb
    except Exception as e:
        print(f"错误: 无法加载数据文件 {DATA_FILE}")
        sys.exit(1)

def find_date_columns(ws):
    """动态查找所有日期列，返回 {datetime: 列索引}"""
    date_cols = {}
    for col in range(1, 500):
        val = ws.cell(1, col).value
        if isinstance(val, datetime):
            date_cols[val] = col
    return date_cols

def get_data(ws, date_cols, target_date):
    """获取指定日期的客流和收入"""
    if target_date not in date_cols:
        return None, None
    col = date_cols[target_date]
    people = ws.cell(9, col).value  # 第9行：客流合计
    income = ws.cell(10, col).value  # 第10行：收入金额
    return people, income

def query_holiday(holiday_type, year='all'):
    """查询节假日数据"""
    wb = load_workbook()
    
    # 节假日配置
    holidays = {
        'qingming': {
            'name': '清明节',
            'dates': {
                2023: [(4, 5), (4, 6), (4, 7)],  # (month, day)
                2024: [(4, 4), (4, 5), (4, 6)],
                2025: [(4, 4), (4, 5), (4, 6)]
            }
        },
        'wuyi': {
            'name': '劳动节',
            'dates': {
                2023: [(4, 29), (4, 30), (5, 1), (5, 2), (5, 3)],  # (month, day)
                2024: [(5, 1), (5, 2), (5, 3), (5, 4), (5, 5)],
                2025: [(5, 1), (5, 2), (5, 3), (5, 4), (5, 5)]
            }
        }
    }
    
    if holiday_type not in holidays:
        print(f"未知节假日类型: {holiday_type}")
        print(f"支持的类型: {', '.join(holidays.keys())}")
        sys.exit(1)
    
    config = holidays[holiday_type]
    years = [year] if year != 'all' else [2023, 2024, 2025]
    
    print(f"\n{'='*60}")
    print(f"  {config['name']}假期历史数据查询")
    print(f"{'='*60}")
    
    results = []
    for y in years:
        if y not in config['dates']:
            continue
            
        ws = wb[f'{y}年']
        date_cols = find_date_columns(ws)
        
        print(f"\n【{y}年{config['name']}】", end=" ")
        print(f"{config['dates'][y][0][0]}月{config['dates'][y][0][1]}日 - {config['dates'][y][-1][0]}月{config['dates'][y][-1][1]}日")
        
        total_people, total_income = 0, 0
        for month, day in config['dates'][y]:
            target_date = datetime(y, month, day)
            people, income = get_data(ws, date_cols, target_date)
            
            if people:
                total_people += people
            if income:
                total_income += income
            
            print(f"  {month}月{day}日: 客流{int(people) if people else 0}, 收入{int(income) if income else 0}")
        
        print(f"  ─────────────────")
        print(f"  合计: 客流{int(total_people)}, 收入{int(total_income):,} ({total_income/10000:.1f}万)")
        
        results.append({
            'year': y,
            'people': total_people,
            'income': total_income
        })
    
    # 排序输出
    if len(results) > 1:
        results.sort(key=lambda x: x['income'], reverse=True)
        print(f"\n{'='*60}")
        print("  收入排名")
        print(f"{'='*60}")
        for i, r in enumerate(results, 1):
            print(f"  {i}. {r['year']}年: {r['income']/10000:.1f}万 ({int(r['people'])}人)")

def query_month(month, year='all'):
    """查询月度数据"""
    wb = load_workbook()
    
    years = [year] if year != 'all' else [2023, 2024, 2025]
    
    print(f"\n{'='*60}")
    print(f"  {month}月历史收入查询")
    print(f"{'='*60}")
    
    results = []
    for y in years:
        ws = wb[f'{y}年']
        date_cols = find_date_columns(ws)
        
        total_income = 0
        days_in_month = 31 if month not in [2, 4, 6, 9, 11] else 30
        if month == 2:
            days_in_month = 29 if (y % 4 == 0 and y % 100 != 0) or (y % 400 == 0) else 28
        
        for day in range(1, days_in_month + 1):
            target_date = datetime(y, month, day)
            people, income = get_data(ws, date_cols, target_date)
            if income:
                total_income += income
        
        print(f"  {y}年{month}月: {int(total_income):,} ({total_income/10000:.1f}万)")
        results.append({'year': y, 'income': total_income})
    
    # 排序
    if len(results) > 1:
        results.sort(key=lambda x: x['income'], reverse=True)
        print(f"\n  → 最高: {results[0]['year']}年 ({results[0]['income']/10000:.1f}万)")

def main():
    parser = argparse.ArgumentParser(description='电影小镇历史数据查询工具')
    parser.add_argument('--type', choices=['qingming', 'wuyi', 'month'], required=True, help='查询类型')
    parser.add_argument('--year', default='all', help='年份 (2023, 2024, 2025, all)')
    parser.add_argument('--month', type=int, help='月份 (与 --type=month 配合使用)')
    
    args = parser.parse_args()
    
    if args.type == 'month' and not args.month:
        print("错误: --type=month 时需要指定 --month")
        sys.exit(1)
    
    year = args.year if args.year == 'all' else int(args.year)
    
    if args.type == 'qingming':
        query_holiday('qingming', year)
    elif args.type == 'wuyi':
        query_holiday('wuyi', year)
    elif args.type == 'month':
        query_month(args.month, year)

if __name__ == '__main__':
    main()
