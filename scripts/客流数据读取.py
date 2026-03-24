#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
电影小镇客流数据读取脚本（防错版）
自动识别表格结构，正确读取各年份数据
"""

from openpyxl import load_workbook
from datetime import datetime
import sys

# 表格路径
HISTORY_FILE = "/Users/tianjinzhan/Desktop/2023-2025年门票销售及客流统计数据表.xlsx"
CURRENT_FILE = "/Users/tianjinzhan/Desktop/2026年电影小镇实际客流.xlsx"

def find_data_row(ws, keyword):
    """根据关键词找到数据所在行"""
    for row in range(1, 20):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and keyword in str(cell_val):
            return row
    return None

def get_year_data(year):
    """获取指定年份1-3月的数据"""
    
    if year == 2026:
        wb = load_workbook(CURRENT_FILE)
        ws = wb.active
    else:
        wb = load_workbook(HISTORY_FILE)
        ws = wb[f'{year}年']
    
    # 找到客流和收入所在的行
    # 2026年：第9行=客流合计，第10行=收入
    # 2023-2025年：第9行=客流合计，第10行=收入
    
    if year == 2026:
        visitors_row = 9
        revenue_row = 10
    else:
        # 查找"门票人数合计"行
        visitors_row = None
        revenue_row = None
        for row in range(1, 15):
            label = ws.cell(row=row, column=1).value
            if label:
                if "门票人数合计" in str(label):
                    visitors_row = row
                if "门票收入金额" in str(label):
                    revenue_row = row
        
        # 如果没找到，使用默认行号
        if not visitors_row:
            visitors_row = 9
        if not revenue_row:
            revenue_row = 10
    
    # 统计1-3月数据
    total_visitors = 0
    total_revenue = 0
    
    for col in range(3, 100):  # 最多检查100列
        date_cell = ws.cell(row=1, column=col).value
        if date_cell and isinstance(date_cell, datetime):
            if date_cell.month in [1, 2, 3]:
                # 读取客流
                visitors = ws.cell(row=visitors_row, column=col).value
                if visitors and isinstance(visitors, (int, float)) and visitors > 0:
                    total_visitors += int(visitors)
                
                # 读取收入
                revenue = ws.cell(row=revenue_row, column=col).value
                if revenue and isinstance(revenue, (int, float)) and revenue > 0:
                    total_revenue += float(revenue)
    
    return {
        "visitors": total_visitors,
        "revenue": total_revenue
    }

def get_month_data(year, month):
    """获取指定年份指定月的数据"""
    
    if year == 2026:
        wb = load_workbook(CURRENT_FILE)
        ws = wb.active
    else:
        wb = load_workbook(HISTORY_FILE)
        ws = wb[f'{year}年']
    
    if year == 2026:
        visitors_row = 9
        revenue_row = 10
    else:
        visitors_row = 9
        revenue_row = 10
    
    total_visitors = 0
    total_revenue = 0
    
    for col in range(3, 100):
        date_cell = ws.cell(row=1, column=col).value
        if date_cell and isinstance(date_cell, datetime):
            if date_cell.month == month:
                visitors = ws.cell(row=visitors_row, column=col).value
                if visitors and isinstance(visitors, (int, float)) and visitors > 0:
                    total_visitors += int(visitors)
                
                revenue = ws.cell(row=revenue_row, column=col).value
                if revenue and isinstance(revenue, (int, float)) and revenue > 0:
                    total_revenue += float(revenue)
    
    return {
        "visitors": total_visitors,
        "revenue": total_revenue
    }

def main():
    # 测试：打印2023-2026年1-3月数据
    print("=" * 50)
    print("电影小镇 Q1 (1-3月) 客流收入数据")
    print("=" * 50)
    
    for year in [2023, 2024, 2025, 2026]:
        data = get_year_data(year)
        print(f"{year}年: 客流={data['visitors']:,}, 收入={data['revenue']:,.0f}")
    
    print("\n2026年各月数据:")
    for month in [1, 2, 3]:
        data = get_month_data(2026, month)
        print(f"  {month}月: 客流={data['visitors']:,}, 收入={data['revenue']:,.0f}")

if __name__ == "__main__":
    main()
