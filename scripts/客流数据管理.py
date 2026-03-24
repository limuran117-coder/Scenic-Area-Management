#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
电影小镇客流数据管理脚本
功能：
1. 录入每日数据
2. 查询历史数据
3. 对比分析
"""

import re
from datetime import datetime
from openpyxl import load_workbook

# 表格路径
CURRENT_FILE = "/Users/tianjinzhan/Desktop/2026年电影小镇实际客流.xlsx"
HISTORY_FILE = "/Users/tianjinzhan/Desktop/2023-2025年门票销售及客流统计数据表.xlsx"

def parse_input(text):
    """解析用户输入，提取日期、收入、客流"""
    
    # 提取日期
    date_pattern = r'(\d{1,2})月(\d{1,2})日'
    date_match = re.search(date_pattern, text)
    
    # 提取收入（支持：21万、217141元、210000等）
    revenue_pattern = r'收?入[：:]*(\d+(?:\.\d+)?)(万|元)?'
    revenue_match = re.search(revenue_pattern, text)
    
    # 提取客流（支持：2833人次、2833人等）
    visitors_pattern = r'客?流[：:]*(\d+)(?:人次|人)?'
    visitors_match = re.search(visitors_pattern, text)
    
    if date_match and revenue_match and visitors_match:
        month = int(date_match.group(1))
        day = int(date_match.group(2))
        
        revenue_val = float(revenue_match.group(1))
        if revenue_match.group(2) == '万':
            revenue_val = revenue_val * 10000
        
        visitors_val = int(visitors_match.group(1))
        
        return {
            "month": month,
            "day": day,
            "revenue": revenue_val,
            "visitors": visitors_val
        }
    
    return None

def write_data(month, day, revenue, visitors):
    """写入数据到2026年表格"""
    
    wb = load_workbook(CURRENT_FILE)
    ws = wb.active
    
    # 找到对应日期的列
    target_col = None
    for col in range(3, 100):
        cell = ws.cell(row=1, column=col)
        if cell.value and isinstance(cell.value, datetime):
            if cell.value.month == month and cell.value.day == day:
                target_col = col
                break
    
    if not target_col:
        return False, f"未找到 {month}月{day}日 对应的列"
    
    # 写入数据
    # 第10行：门票收入金额
    ws.cell(row=10, column=target_col, value=revenue)
    # 第9行：门票人数合计
    ws.cell(row=9, column=target_col, value=visitors)
    
    wb.save(CURRENT_FILE)
    
    return True, f"已录入 {month}月{day}日: 收入{revenue}元, 客流{visitors}人次"

def get_year_q1_data(year):
    """获取指定年份Q1数据"""
    
    if year == 2026:
        wb = load_workbook(CURRENT_FILE)
        ws = wb.active
        visitors_row = 9
        revenue_row = 10
    else:
        wb = load_workbook(HISTORY_FILE)
        ws = wb[f'{year}年']
        visitors_row = 9
        revenue_row = 10
    
    total_visitors = 0
    total_revenue = 0
    
    for col in range(3, 100):
        date_cell = ws.cell(row=1, column=col).value
        if date_cell and isinstance(date_cell, datetime):
            if date_cell.month in [1, 2, 3]:
                v = ws.cell(row=visitors_row, column=col).value
                r = ws.cell(row=revenue_row, column=col).value
                if v and isinstance(v, (int, float)) and v > 0:
                    total_visitors += int(v)
                if r and isinstance(r, (int, float)) and r > 0:
                    total_revenue += float(r)
    
    return {"visitors": total_visitors, "revenue": total_revenue}

def compare_with_history():
    """对比历史数据"""
    
    # 2026年Q1
    data_26 = get_year_q1_data(2026)
    
    # 2025年Q1
    data_25 = get_year_q1_data(2025)
    
    # 计算同比
    vis_change = (data_26['visitors'] / data_25['visitors'] - 1) * 100 if data_25['visitors'] > 0 else 0
    rev_change = (data_26['revenue'] / data_25['revenue'] - 1) * 100 if data_25['revenue'] > 0 else 0
    
    return {
        "2026": data_26,
        "2025": data_25,
        "vis_change": vis_change,
        "rev_change": rev_change
    }

# 测试
if __name__ == "__main__":
    # 测试解析
    test_text = "3月19日，收入217141元，客流2833人次"
    result = parse_input(test_text)
    print("解析测试:", result)
    
    # 测试对比
    comp = compare_with_history()
    print(f"\n2026年Q1: 客流{comp['2026']['visitors']:,}, 收入{comp['2026']['revenue']:,.0f}")
    print(f"2025年Q1: 客流{comp['2025']['visitors']:,}, 收入{comp['2025']['revenue']:,.0f}")
    print(f"同比: 客流{comp['vis_change']:+.1f}%, 收入{comp['rev_change']:+.1f}%")
