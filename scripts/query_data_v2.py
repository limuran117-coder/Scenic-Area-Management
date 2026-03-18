#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
电影小镇历史数据查询工具 v2.0
- 支持多种节假日查询
- 自动校验数据
- 防止查询失误
"""

import openpyxl
from datetime import datetime
import argparse
import sys
import json
import os

# 配置
DATA_FILE = '2023-2025年门票销售及客流统计数据表.xlsx'
OUTPUT_DIR = 'data_summary'

# 节假日配置（官方放假日期）
HOLIDAYS_CONFIG = {
    '春节': {
        'dates': {
            2023: [(1, 22), (1, 23), (1, 24), (1, 25), (1, 26), (1, 27), (1, 28)],  # 1月22日-28日
            2024: [(2, 10), (2, 11), (2, 12), (2, 13), (2, 14), (2, 15), (2, 16), (2, 17)],  # 2月10日-17日
            2025: [(1, 28), (1, 29), (1, 30), (1, 31), (2, 1), (2, 2), (2, 3), (2, 4)],  # 1月28日-2月4日
        }
    },
    '清明节': {
        'dates': {
            2023: [(4, 5), (4, 6), (4, 7)],
            2024: [(4, 4), (4, 5), (4, 6)],
            2025: [(4, 4), (4, 5), (4, 6)],
        }
    },
    '劳动节': {
        'dates': {
            2023: [(4, 29), (4, 30), (5, 1), (5, 2), (5, 3)],
            2024: [(5, 1), (5, 2), (5, 3), (5, 4), (5, 5)],
            2025: [(5, 1), (5, 2), (5, 3), (5, 4), (5, 5)],
        }
    },
    '端午节': {
        'dates': {
            2023: [(6, 22), (6, 23), (6, 24)],  # 6月22日-24日
            2024: [(6, 8), (6, 9), (6, 10)],  # 6月8日-10日
            2025: [(5, 31), (6, 1), (6, 2)],  # 5月31日-6月2日
        }
    },
    '中秋节': {
        'dates': {
            2023: [(9, 29), (9, 30), (10, 1)],  # 9月29日-10月1日
            2024: [(9, 15), (9, 16), (9, 17)],  # 9月15日-17日
            2025: [(10, 6), (10, 7), (10, 8)],  # 10月6日-8日
        }
    },
    '国庆节': {
        'dates': {
            2023: [(9, 29), (9, 30), (10, 1), (10, 2), (10, 3), (10, 4), (10, 5), (10, 6)],  # 9月29日-10月6日
            2024: [(10, 1), (10, 2), (10, 3), (10, 4), (10, 5), (10, 6), (10, 7)],  # 10月1日-7日
            2025: [(10, 1), (10, 2), (10, 3), (10, 4), (10, 5), (10, 6), (10, 7), (10, 8)],  # 10月1日-8日
        }
    },
}


class DataQueryTool:
    """数据查询工具类"""
    
    def __init__(self, data_file=DATA_FILE):
        self.data_file = data_file
        self.wb = None
        self.date_columns_cache = {}
        
    def load_data(self):
        """加载数据文件"""
        try:
            self.wb = openpyxl.load_workbook(self.data_file)
            print(f"✓ 数据文件加载成功: {self.data_file}")
            return True
        except Exception as e:
            print(f"✗ 加载数据失败: {e}")
            return False
    
    def get_date_columns(self, year):
        """获取某年的日期列缓存"""
        if year not in self.date_columns_cache:
            ws = self.wb[f'{year}年']
            date_cols = {}
            for col in range(1, 500):
                val = ws.cell(1, col).value
                if isinstance(val, datetime):
                    date_cols[val] = col
            self.date_columns_cache[year] = date_cols
        return self.date_columns_cache[year]
    
    def get_daily_data(self, year, month, day):
        """获取单日数据"""
        date_cols = self.get_date_columns(year)
        target_date = datetime(year, month, day)
        
        if target_date not in date_cols:
            return None, None
            
        col = date_cols[target_date]
        ws = self.wb[f'{year}年']
        people = ws.cell(9, col).value  # 客流
        income = ws.cell(10, col).value  # 收入
        return people, income
    
    def query_holiday(self, holiday_name, year=None):
        """查询节假日数据"""
        if holiday_name not in HOLIDAYS_CONFIG:
            print(f"✗ 未知节假日: {holiday_name}")
            print(f"支持的节假日: {', '.join(HOLIDAYS_CONFIG.keys())}")
            return None
            
        config = HOLIDAYS_CONFIG[holiday_name]
        years = [year] if year else sorted(config['dates'].keys())
        
        results = []
        
        print(f"\n{'='*60}")
        print(f"  【{holiday_name}】历史数据查询")
        print(f"{'='*60}")
        
        for y in years:
            if y not in config['dates']:
                continue
                
            dates = config['dates'][y]
            print(f"\n【{y}年{holiday_name}】{dates[0][0]}月{dates[0][1]}日 - {dates[-1][0]}月{dates[-1][1]}日")
            
            total_people, total_income = 0, 0
            daily_data = []
            
            for month, day in dates:
                people, income = self.get_daily_data(y, month, day)
                people = people if people else 0
                income = income if income else 0
                
                total_people += people
                total_income += income
                daily_data.append({'date': f'{month}月{day}日', 'people': people, 'income': income})
                
                print(f"  {month}月{day}日: 客流{int(people):,}, 收入{int(income):,}")
            
            print(f"  ─────────────────")
            print(f"  合计: 客流{int(total_people):,}, 收入{int(total_income):,} ({total_income/10000:.1f}万)")
            
            results.append({
                'year': y,
                'holiday': holiday_name,
                'dates': f'{dates[0][0]}月{dates[0][1]}日-{dates[-1][0]}月{dates[-1][1]}日',
                'days': len(dates),
                'total_people': total_people,
                'total_income': total_income,
                'avg_people': total_people / len(dates),
                'avg_income': total_income / len(dates),
                'daily': daily_data
            })
        
        # 排序
        if len(results) > 1:
            results.sort(key=lambda x: x['total_income'], reverse=True)
            print(f"\n{'='*60}")
            print("  收入排名")
            print(f"{'='*60}")
            for i, r in enumerate(results, 1):
                print(f"  {i}. {r['year']}年: {r['total_income']/10000:.1f}万 ({int(r['total_people']):,}人)")
        
        return results
    
    def query_month(self, month, year=None):
        """查询月度数据"""
        years = [year] if year else [2023, 2024, 2025]
        
        results = []
        
        print(f"\n{'='*60}")
        print(f"  【{month}月】历史数据查询")
        print(f"{'='*60}")
        
        for y in years:
            total_income = 0
            total_people = 0
            
            # 计算该月天数
            if month == 2:
                days = 29 if (y % 4 == 0 and y % 100 != 0) or (y % 400 == 0) else 28
            elif month in [1, 3, 5, 7, 8, 10, 12]:
                days = 31
            else:
                days = 30
            
            for day in range(1, days + 1):
                people, income = self.get_daily_data(y, month, day)
                if people:
                    total_people += people
                if income:
                    total_income += income
            
            print(f"  {y}年{month}月: 客流{int(total_people):,}, 收入{int(total_income):,} ({total_income/10000:.1f}万)")
            
            results.append({
                'year': y,
                'month': month,
                'total_people': total_people,
                'total_income': total_income
            })
        
        if len(results) > 1:
            results.sort(key=lambda x: x['total_income'], reverse=True)
            print(f"\n  → 最高: {results[0]['year']}年 ({results[0]['total_income']/10000:.1f}万)")
        
        return results
    
    def validate_data(self):
        """数据校验"""
        print(f"\n{'='*60}")
        print("  数据校验")
        print(f"{'='*60}")
        
        errors = []
        
        # 检查数据文件
        for year in [2023, 2024, 2025]:
            sheet_name = f'{year}年'
            if sheet_name not in self.wb.sheetnames:
                errors.append(f"缺少工作表: {sheet_name}")
        
        if errors:
            print("✗ 校验失败:")
            for e in errors:
                print(f"  - {e}")
            return False
        
        print("✓ 数据文件校验通过")
        
        # 抽样检查
        print("\n抽样检查（清明节4月5日）:")
        for year in [2023, 2024, 2025]:
            people, income = self.get_daily_data(year, 4, 5)
            print(f"  {year}年4月5日: 客流{people}, 收入{income}")
        
        return True
    
    def generate_all_summary(self):
        """生成全年数据汇总"""
        print(f"\n{'='*60}")
        print("  生成全年数据汇总")
        print(f"{'='*60}")
        
        all_data = {
            'holidays': {},
            'months': {}
        }
        
        # 查询所有节假日
        for holiday_name in HOLIDAYS_CONFIG:
            results = self.query_holiday(holiday_name)
            if results:
                all_data['holidays'][holiday_name] = results
        
        # 查询所有月份
        for month in range(1, 13):
            results = self.query_month(month)
            if results:
                all_data['months'][month] = results
        
        return all_data


def main():
    parser = argparse.ArgumentParser(description='电影小镇历史数据查询工具 v2.0')
    parser.add_argument('--type', '-t', choices=['holiday', 'month', 'validate', 'all'], 
                       default='validate', help='查询类型')
    parser.add_argument('--name', '-n', help='节假日名称 (与 --type=holiday 配合)')
    parser.add_argument('--month', '-m', type=int, help='月份 (与 --type=month 配合)')
    parser.add_argument('--year', '-y', type=int, help='年份')
    
    args = parser.parse_args()
    
    tool = DataQueryTool()
    
    if not tool.load_data():
        sys.exit(1)
    
    if args.type == 'validate':
        tool.validate_data()
        
    elif args.type == 'holiday':
        if not args.name:
            print("请指定节假日名称: --name 清明节")
            print(f"支持的节假日: {', '.join(HOLIDAYS_CONFIG.keys())}")
            sys.exit(1)
        tool.query_holiday(args.name, args.year)
        
    elif args.type == 'month':
        if not args.month:
            print("请指定月份: --month 5")
            sys.exit(1)
        tool.query_month(args.month, args.year)
        
    elif args.type == 'all':
        tool.generate_all_summary()


if __name__ == '__main__':
    main()
